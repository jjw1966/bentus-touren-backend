from fastapi import FastAPI
from fastapi.middleware.cors import CORSMiddleware
from openpyxl import load_workbook
from datetime import datetime
import requests
from io import BytesIO

app = FastAPI()

# 🟩 Tillåt frontend på Render
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# 🟩 Google Drive-länk till Excel
GOOGLE_DRIVE_URL = "https://docs.google.com/spreadsheets/d/1a6gNhvSLO6kEBzZbWM2He8Zn0vblFNcgrE4iGr-b_ko/export?format=xlsx

# 🟩 Health check (Render kräver denna)
@app.get("/health")
def health():
    return {"status": "ok"}


# 🟩 Funktion för att hämta Excel från Drive med fallback
def load_excel():
    try:
        response = requests.get(GOOGLE_DRIVE_URL, timeout=10)
        response.raise_for_status()
        return load_workbook(BytesIO(response.content), data_only=True)
    except Exception as e:
        print(f"⚠️ Kunde inte hämta från Drive: {e}")
        try:
            return load_workbook("BentusTouren.xlsx", data_only=True)
        except Exception as e2:
            print(f"⚠️ Kunde inte läsa lokal fil heller: {e2}")
            raise RuntimeError("Excel-filen kunde inte laddas från Drive eller lokalt.")


# 🟩 Dynamisk tabell-läsare
def read_table(sheet, header_text):
    header_row = None
    for i, row in enumerate(sheet.iter_rows(values_only=True)):
        if row and isinstance(row[0], str) and row[0].strip().startswith(header_text):
            header_row = i + 1
            break

    if header_row is None:
        return []

    columns = [c for c in sheet[header_row] if c]

    data = []
    for r in sheet.iter_rows(min_row=header_row + 1, values_only=True):
        if not any(r):
            break
        entry = dict(zip(columns, r))
        data.append(entry)

    return data


# 🟩 Dashboard-endpoint
@app.get("/dashboard")
def get_dashboard():
    wb = load_excel()
    sheet = wb["Dashboard"]

    dashboard = {
        "topp5": read_table(sheet, "Topp"),
        "nh": read_table(sheet, "Närmast"),
        "ld": read_table(sheet, "Längsta"),
        "spelade": read_table(sheet, "Spelade"),
        "vinster": read_table(sheet, "Deltävlingsvinster"),
        "landskamper": read_table(sheet, "Landskamper"),
        "deltavlingar": read_table(sheet, "Deltävlingar"),
    }

    for d in dashboard.get("deltavlingar", []):
        if "Datum" in d and isinstance(d["Datum"], datetime):
            d["Datum"] = d["Datum"].strftime("%Y-%m-%d")

    return dashboard


# 🟩 Tourställning-endpoint
@app.get("/tourstallning")
def get_tourstallning():
    wb = load_excel()
    sheet = wb["Tourställning"]

    tour = {
        "tourstallning": read_table(sheet, "Tourställning"),
        "poang": read_table(sheet, "Poäng"),
        "rundor": read_table(sheet, "Rundor"),
        "vinster": read_table(sheet, "Vinster"),
        "statistik": read_table(sheet, "Statistik"),
    }

    return tour


# 🟩 Root-endpoint
@app.get("/")
def root():
    return {"status": "Bentus Touren backend aktiv"}
